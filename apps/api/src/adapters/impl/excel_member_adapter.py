from __future__ import annotations

import time
from pathlib import Path

import openpyxl

from src.models.member_verify import MemberVerify
from src.services.member.port import MemberPort

_REQUIRED_COLUMNS = ("이름", "이메일", "팀")
_CACHE_TTL_SECONDS = 300.0  # NFR-12: 회원 목록은 5분 메모리 캐시


class ExcelMemberAdapter(MemberPort):
    """로컬 Members.xlsx를 읽어 회원을 검증한다 (ADR-001 / PRD v4.0).

    컬럼은 헤더명(이름/이메일/팀)으로 매핑하므로 열 순서에 의존하지 않는다.
    이메일은 strip + 소문자로 정규화하며, 미등재 이메일은 None을 반환한다.
    파일은 매 조회마다 다시 읽지 않고 TTL(기본 5분) 동안 메모리에 캐시한다.
    """

    def __init__(
        self,
        xlsx_path: str | Path,
        *,
        cache_ttl_seconds: float = _CACHE_TTL_SECONDS,
    ) -> None:
        self._path = Path(xlsx_path)
        self._cache_ttl = cache_ttl_seconds
        self._members: dict[str, MemberVerify] = {}
        self._loaded_at: float | None = None

    def verify(self, email: str) -> MemberVerify | None:
        self._ensure_loaded()
        return self._members.get(self._normalize(email))

    def add(self, member: MemberVerify) -> None:
        # 헤더 컬럼 위치를 찾아 한 행 append 후 저장. 캐시 무효화로 다음 verify가 재로딩.
        workbook = openpyxl.load_workbook(self._path)
        try:
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError("Members.xlsx 활성 시트가 없습니다.")
            header = [str(c.value).strip() if c.value is not None else "" for c in worksheet[1]]
            row: list[str] = [""] * len(header)
            row[header.index("이름")] = member.name
            row[header.index("이메일")] = member.email
            row[header.index("팀")] = member.team
            worksheet.append(row)
            workbook.save(self._path)
        finally:
            workbook.close()
        self._loaded_at = None

    @staticmethod
    def _normalize(email: str) -> str:
        return email.strip().lower()

    def _ensure_loaded(self) -> None:
        fresh = (
            self._loaded_at is not None and (time.monotonic() - self._loaded_at) < self._cache_ttl
        )
        if fresh:
            return
        self._members = self._load()
        self._loaded_at = time.monotonic()

    def _load(self) -> dict[str, MemberVerify]:
        if not self._path.exists():
            raise FileNotFoundError(f"Members 파일을 찾을 수 없습니다: {self._path}")

        workbook = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                return {}
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                return {}

            col = {str(name).strip(): i for i, name in enumerate(header) if name is not None}
            missing = [c for c in _REQUIRED_COLUMNS if c not in col]
            if missing:
                raise ValueError(f"Members.xlsx 필수 컬럼 누락: {missing}")
            i_name, i_email, i_team = col["이름"], col["이메일"], col["팀"]

            members: dict[str, MemberVerify] = {}
            for row in rows:
                name, email, team = row[i_name], row[i_email], row[i_team]
                if not (name and email and team):
                    continue
                key = self._normalize(str(email))
                members[key] = MemberVerify(
                    email=key, name=str(name).strip(), team=str(team).strip()
                )
            return members
        finally:
            workbook.close()
